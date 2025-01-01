from math import fabs
from openpyxl import load_workbook


import argparse
import os
import glob
import jinja2
import sys
import logging


def setup_logging():
    logging.basicConfig(level=logging.ERROR, format=" \033[91m%(message) In %(xlsxName)  s\033[0m ")

def PrintError( xlsxName, message):
    logging.error(xlsxName, message)
    sys.exit(1)


class Member:
    def __init__(self, Name, valueType, default):
        self.Name = Name
        self.name = Name[0].lower() + Name[1:]
        self.valueType = valueType
        self.default = default


infoManagerNames = []
infoManagerNameAndPaths = []


def process_files_in_directory( file_path, destPath, defaultDestPath ):
    
    data = []
    rows = []

##### xlsx parser
    workbook = load_workbook(file_path)
    
    for sheetIndex, sheetname in enumerate(workbook.sheetnames):
        if 0 < sheetIndex:
            continue

        # sheet
        infoManagerNames.append(sheetname)
        path = file_path
        if path.startswith("../../Data\\"):
            path = path[len("../../Data\\"):]
    
        if path.endswith(".xlsx"):
            path = path[:-len(".xlsx")]
    
        path = path.replace("\\", "/")

        infoManagerNameAndPaths.append(path)


        sheet = workbook[sheetname]

        for rowIndex, row in enumerate(sheet.iter_rows(values_only=True)):
            rowList = list(row)

            data.append(rowList)

            if ( 3 < rowIndex ):
                if not rowList[1] == 1:   ## Check Disable
                    rows.append(list(row))

        # print
        #for row in data:
        #    print(row)

    for row in rows:
        for i in range(len(row)):
            if row[i] is None:
                row[i] = 0;

            if isinstance(row[i], str):
                dataType = data[3][i]

                if dataType.startswith("E"):
                    row[i] = "Protocol::" +dataType+ "::"+ row[i]
                if dataType == "AtString":
                    row[i] = "\"" + row[i] + "\""


#### delete Disable Column
    for row in rows:
        for i in range(len(row)):
            b = 0;

            if i == 1:
                del row[i]
                b = 1;

            if b == 1:
                break



    if data[0][0] != "Id":
        PrintError(file_path, "No Have 'Id' Column. [0][0]")

    if data[0][1] != "Disable":
        PrintError(file_path, "No Have 'Disable' Column. [0][1]")

    if not data[2][0].startswith('*'):
        PrintError(file_path, "No Have PRYKEY in 'Id'")


##### set member
    memberList = []

    for colIndex, element in enumerate(data[0]):
        if colIndex == 1: # DisableColumn
            continue
    
        memberNAME    = data[0][colIndex]

        memberType    = data[3][colIndex]
        if memberType.startswith("E"):
            memberType = "Protocol::" + memberType

        memberDefault =  memberType + "( 0 )"

        if memberType == "AtString":
            memberDefault = "\"\""
        if memberType.startswith("E"):
            memberDefault = memberType + "::Max"

        memberList.append(Member(memberNAME, memberType, memberDefault))


##### Make Dictory ( isNoHave )
    subString = "../../Data\\"
    index = file_path.find(subString) + len(subString)
    result = file_path[index:]
    
    index = result.find("\\")
    
    if index != -1:
        result = result[:index]
    
        destPath = destPath + '/' + result + '/'
        if not os.path.exists(destPath):
            os.makedirs(destPath)
###########
    
    file_loader = jinja2.FileSystemLoader('Templates')
    env = jinja2.Environment(loader=file_loader)
    
##### Make Template ( ForceMake )
    template = env.get_template('InfoTemplate.h')
    output = template.render(ClassName=sheetname, memberList=memberList)
    
    f = open(destPath + sheetname+'InfoTemplate.h', 'w+')
    f.write(output)
    f.close()
    
    template = env.get_template('InfoTemplate.cpp')
    output = template.render(ClassName=sheetname, memberList=memberList)
    
    f = open(destPath +sheetname+'InfoTemplate.cpp', 'w+')
    f.write(output)
    f.close()
###########
    
##### Make Info ( No Have? Make! )
    template = env.get_template('Info.h')
    output = template.render(ClassName=sheetname, memberList=memberList)
    
    if not os.path.exists(destPath +sheetname+'Info.h'):
        f = open(destPath +sheetname+'Info.h', 'w+')
        f.write(output)
        f.close()
    
    template = env.get_template('Info.cpp')
    output = template.render(ClassName=sheetname, memberList=memberList)
    
    if not os.path.exists(destPath +sheetname+'Info.cpp'):
        f = open(destPath +sheetname+'Info.cpp', 'w+')
        f.write(output)
        f.close()
###########


    dataPryKeyType = data[3][0]
    dataPryKey     = data[0][0]
    dataPryKey     = dataPryKey[0].lower() + dataPryKey[1:]


##### Make InfoManagerTemplate ( ForceMake )
    template = env.get_template('InfoManagerTemplate.h')
    output = template.render(ClassName=sheetname, memberList=memberList, KeyType = dataPryKeyType, KeyName = dataPryKey, rows = rows)
    
    f = open(destPath + sheetname+'InfoManagerTemplate.h', 'w+')
    f.write(output)
    f.close()
    
    template = env.get_template('InfoManagerTemplate.cpp')
    output = template.render(ClassName=sheetname, memberList=memberList, KeyType = dataPryKeyType, KeyName = dataPryKey, rows = rows)
    
    f = open(destPath +sheetname+'InfoManagerTemplate.cpp', 'w+')
    f.write(output)
    f.close()
###########


##### Make InfoManager ( No Have? Make! )
    template = env.get_template('InfoManager.h')
    output = template.render(ClassName=sheetname)
    
    if not os.path.exists(destPath +sheetname+'InfoManager.h'):
        f = open(destPath +sheetname+'InfoManager.h', 'w+')
        f.write(output)
        f.close()
    
    template = env.get_template('InfoManager.cpp')
    output = template.render(ClassName=sheetname)
    
    if not os.path.exists(destPath +sheetname+'InfoManager.cpp'):
        f = open(destPath +sheetname+'InfoManager.cpp', 'w+')
        f.write(output)
        f.close()
###########


##### Make InfoManagers ( ForceMake )
    template = env.get_template('InfoManagers.h')
    output = template.render(InfoManagerNameAndPaths = infoManagerNameAndPaths, InfoManagerNames = infoManagerNames)
    
    f = open(defaultDestPath +'InfoManagers.h', 'w+')
    f.write(output)
    f.close()
###########


def main():
    arg_parser = argparse.ArgumentParser(description = 'DataGenerator')
    arg_parser.add_argument('--sourcePath', type=str, default='../../Data', help='data dir path')
    arg_parser.add_argument('--destPath', type=str, default='../../GameServer/Data/', help='data dir path')
    arg_parser.add_argument('--defaultDestPath', type=str, default='../../GameServer/Data/', help='data default dir path')
    args = arg_parser.parse_args()

    for root, dirs, files in os.walk(args.sourcePath):
        for file in files:
            if file.endswith(".xlsx"):

                file_path = os.path.join(root, file)

                process_files_in_directory(file_path, args.destPath, args.defaultDestPath)
                
    return



if __name__ == '__main__':
	main()